from operator import attrgetter
from typing import Dict, List, Iterable, Optional

import attr
from openpyxl import Workbook
from openpyxl.conftest import Worksheet
from openpyxl.utils import get_column_letter

from .models import Style, CellWithSize
from .utils import str_cell_range, style_and_merge_cell_range, style_single_cell


def xlsx_from_json(json_data: Dict, default_style: Style = None) -> Workbook:
    default_style = default_style or Style()

    wb = Workbook()
    sheet = wb.active

    start_row = json_data.get("start_row", 1)
    start_column = json_data.get("start_column", 1)

    filler = RowFiller(
        sheet, start_column, start_row, default_style, json_data.get("number_format")
    )
    row_positions = filler.fill(json_data.get("rows", []))

    adjuster = Adjuster(sheet)
    adjuster.adjust_columns(json_data.get("column_widths", []))
    adjuster.adjust_rows(json_data.get("rows", []), row_positions)

    return wb


@attr.s(auto_attribs=True)
class RowFiller:
    sheet: Worksheet
    start_column: int
    start_row: int
    default_style: Style
    number_format: Optional[str] = None

    def fill(self, rows_data: List[Dict]) -> Iterable[int]:
        current_row = self.start_row

        for row_data in rows_data:
            current_row += row_data.get("rows_shift", 0)
            start_column = self.start_column + row_data.get("columns_shift", 0)

            cells = self._fill_row(current_row, start_column, row_data.get("cells", []))
            row_height = max(map(attrgetter("height"), cells), default=1)

            yield current_row
            current_row += row_height

    def _fill_row(self, row: int, column: int, cells_data: List[Dict]) -> Iterable[CellWithSize]:
        current_column = column

        for cell_data in cells_data:
            cell = self._create_cell(row, current_column, cell_data)
            current_column += cell.width
            yield cell

    def _create_cell(self, row: int, column: int, cell_data: Dict) -> CellWithSize:
        cell = self.sheet.cell(row, column)

        value = cell_data["value"]
        cell.value = value

        if isinstance(value, float) and self.number_format:
            cell.number_format = self.number_format

        width = cell_data.get("width", 1)
        height = cell_data.get("height", 1)

        style = Style.from_json(cell_data.get("style", {}), self.default_style)

        if width == 1 and height == 1:
            style_single_cell(cell, style)
        else:
            cell_range = str_cell_range(column, row, column + width - 1, row + height - 1)
            style_and_merge_cell_range(self.sheet, cell_range, style)

        return CellWithSize(cell, width, height)


@attr.s(auto_attribs=True)
class Adjuster:
    sheet: Worksheet

    def adjust_columns(self, columns: Iterable[Dict]) -> None:
        for column_data in columns:
            if "column_number" in column_data:
                column_letter = get_column_letter(column_data["column_number"])
            elif "column_letter" in column_data:
                column_letter = column_data["column_letter"]
            else:
                raise ValueError("No column provided.")

            self.sheet.column_dimensions[column_letter].width = column_data["width"]

    def adjust_rows(self, rows: Iterable[Dict], row_positions: Iterable[int]) -> None:
        for row_data, row_position in zip(rows, row_positions):
            if "row_height" in row_data:
                self.sheet.row_dimensions[row_position].height = row_data["row_height"]
