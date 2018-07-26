import attr
from openpyxl.cell import Cell
from openpyxl.conftest import Worksheet
from openpyxl.styles import Border
from openpyxl.utils import get_column_letter

from .models import Style


def str_cell_range(start_column: int, start_row: int, end_column: int, end_row: int) -> str:
    from_column = get_column_letter(start_column)
    to_column = get_column_letter(end_column)
    return f"{from_column}{start_row}:{to_column}{end_row}"


def style_single_cell(cell: Cell, style: Style) -> None:
    for attrib, value in attr.asdict(style).items():
        if value:
            setattr(cell, attrib, value)


def style_and_merge_cell_range(ws: Worksheet, cell_range: str, style: Style) -> None:
    """
    Source:
    https://openpyxl.readthedocs.io/en/2.5/styles.html#styling-merged-cells
    """
    top = Border(top=style.border.top)
    left = Border(left=style.border.left)
    right = Border(right=style.border.right)
    bottom = Border(bottom=style.border.bottom)

    first_cell = ws[cell_range.split(":")[0]]
    if style.alignment:
        ws.merge_cells(cell_range)
        first_cell.alignment = style.alignment

    if style.font:
        first_cell.font = style.font

    rows = ws[cell_range]
    for cell in rows[0]:
        cell.border = cell.border + top
    for cell in rows[-1]:
        cell.border = cell.border + bottom

    for row in rows:
        l = row[0]
        r = row[-1]
        l.border = l.border + left
        r.border = r.border + right
        if style.fill:
            for c in row:
                c.fill = style.fill
