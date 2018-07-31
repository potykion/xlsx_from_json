import pytest
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from xlsx_from_json import xlsx_from_json, Style


@pytest.fixture()
def default_style():
    return Style(font=Font(bold=True))


@pytest.fixture()
def json_data_with_start_row_and_multiple_cells():
    return {
        "rows": [
            {
                "cells": [
                    {
                        "value": "1x2",
                        "height": 2
                    }
                ]
            },
            {
                "cells": [
                    {
                        "value": "2x1",
                        "width": 2
                    },
                    {
                        "value": "1x1"
                    }
                ]
            }
        ],
        "start_row": 2
    }


def test_sheet_has_values():
    wb = xlsx_from_json({"rows": [{"cells": [{"value": "op"}]}]})
    sheet = wb.active
    assert sheet.cell(1, 1).value == "op"


def test_cell_column_equals_to_start_column():
    wb = xlsx_from_json({"rows": [{"cells": [{"value": "op"}]}], "start_column": 3})
    sheet = wb.active
    assert sheet.cell(1, 3).value == "op"


def test_cell_has_font():
    wb = xlsx_from_json({
        "rows": [
            {
                "cells": [
                    {
                        "value": "op",
                        "style": {"font": {"size": 12, "name": "Times New Roman"}}
                    }
                ]
            }
        ]
    })
    sheet = wb.active
    cell = sheet.cell(1, 1)
    assert cell.font.name == "Times New Roman"
    assert cell.font.size == 12


def test_cell_has_font_bold_from_default_style(default_style):
    wb = xlsx_from_json({"rows": [{"cells": [{"value": "op"}]}]}, default_style)
    sheet = wb.active
    cell = sheet.cell(1, 1)
    assert cell.font.bold


def test_cell_has_border():
    wb = xlsx_from_json({
        "rows": [
            {
                "cells": [
                    {
                        "value": "op",
                        "style": {"border": {"bottom": {"border_style": "medium", "color": "00000000"}}}
                    }
                ]
            }
        ]
    })
    sheet = wb.active
    cell = sheet.cell(1, 1)
    assert cell.border.bottom.border_style == "medium"
    assert cell.border.bottom.color.rgb == "00000000"


def test_cell_has_alignment():
    wb = xlsx_from_json({
        "rows": [
            {
                "cells": [
                    {
                        "value": "op",
                        "style": {"alignment": {"horizontal": "center"}}
                    }
                ]
            }
        ]
    })
    sheet = wb.active
    cell = sheet.cell(1, 1)
    assert cell.alignment.horizontal == "center"


def test_cell_has_fill():
    wb = xlsx_from_json({
        "rows": [
            {
                "cells": [
                    {
                        "value": "op",
                        "style": {"fill": {"patternType": "gray125"}}
                    }
                ]
            }
        ]
    })
    sheet = wb.active
    cell = sheet.cell(1, 1)
    assert cell.fill.patternType == "gray125"


def test_sized_cell_is_rendered_as_merged_cells_and_style_set():
    workbook: Workbook = xlsx_from_json({
        "rows": [
            {
                "cells": [
                    {
                        "value": "Sample text",
                        "width": 5,
                        "height": 2,
                        "style": {
                            "font": {
                                "name": "Times New Roman",
                                "size": 12
                            },
                            "border": {
                                "bottom": {
                                    "border_style": "medium",
                                    "color": "FFFFFFFF"
                                }
                            }
                        }
                    }
                ]
            }
        ]
    })
    sheet = workbook.active
    assert sheet.cell(2, 5).border.bottom.border_style == "medium"


def test_sheet_fill_starts_with_start_row(json_data_with_start_row_and_multiple_cells):
    workbook: Workbook = xlsx_from_json(json_data_with_start_row_and_multiple_cells)
    sheet = workbook.active
    assert sheet.cell(row=2, column=1).value == "1x2"


def test_sheet_has_two_rows(json_data_with_start_row_and_multiple_cells):
    workbook: Workbook = xlsx_from_json(json_data_with_start_row_and_multiple_cells)
    sheet = workbook.active
    assert sheet.cell(row=2, column=1).value == "1x2"
    assert sheet.cell(row=4, column=1).value == "2x1"


def test_sheet_has_two_columns(json_data_with_start_row_and_multiple_cells):
    workbook: Workbook = xlsx_from_json(json_data_with_start_row_and_multiple_cells)
    sheet = workbook.active
    assert sheet.cell(row=4, column=1).value == "2x1"
    assert sheet.cell(row=4, column=3).value == "1x1"


def test_empty_row_render():
    workbook: Workbook = xlsx_from_json({
        "rows": [
            {"cells": []},
            {"cells": [{"value": "op"}]},
            {},
            {"cells": [{"value": "op"}]},
        ]
    })
    sheet = workbook.active
    assert sheet.cell(2, 1).value == "op"
    assert sheet.cell(4, 1).value == "op"


def test_row_and_column_shift():
    workbook: Workbook = xlsx_from_json({
        "rows": [
            {"cells": [{"value": "op"}], "rows_shift": 2, "columns_shift": 3}
        ]
    })
    sheet = workbook.active
    assert sheet.cell(3, 4).value == "op"


def test_rows_and_columns_shift_for_cell():
    workbook: Workbook = xlsx_from_json({
        "rows": [
            {"cells": [{"value": "op", "rows_shift": 2, "columns_shift": 3}]},
            {"cells": [{"value": "op"}]},
        ]
    })
    sheet = workbook.active
    assert sheet.cell(row=3, column=4).value == "op"
    assert sheet.cell(row=4, column=1).value == "op"


def test_rows_and_columns_shift_for_cell_with_height_ignore():
    workbook: Workbook = xlsx_from_json({
        "rows": [
            {"cells": [{"value": "op", "rows_shift": 2, "columns_shift": 3, "ignore_height": True}]},
            {"cells": [{"value": "op"}]},
        ]
    })
    sheet = workbook.active
    assert sheet.cell(row=3, column=4).value == "op"
    assert sheet.cell(row=2, column=1).value == "op"


def test_row_and_column_sizing():
    workbook: Workbook = xlsx_from_json({
        "rows": [
            {"cells": [{"value": "op"}], "row_height": 10},
        ],
        "column_widths": [
            {
                "width": 10,
                "column_number": 1
            },
            {
                "width": 20,
                "column_letter": "B"
            }
        ]
    })
    sheet = workbook.active
    assert sheet.row_dimensions[1].height == 10
    assert sheet.column_dimensions[get_column_letter(1)].width == 10
    assert sheet.column_dimensions["B"].width == 20


def test_float_formatting():
    workbook: Workbook = xlsx_from_json({
        "rows": [
            {"cells": [{"value": 1.256}]},
        ],
        "number_format": "0.00"
    })
    sheet = workbook.active
    assert sheet.cell(1, 1).number_format == "0.00"
